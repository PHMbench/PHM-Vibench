"""Download and validate versioned PHMFactory dataset bundles.

The provider layer is intentionally separate from the protected Data Factory. It
materializes a local directory containing files that the existing runtime already
understands, then returns that directory without changing reader or cache logic.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
import shutil
import subprocess
import tempfile
from typing import Any, Callable, Mapping, Sequence

import yaml


class BundleValidationError(ValueError):
    """Raised when a downloaded bundle violates its declared data contract."""


@dataclass(frozen=True)
class BundleFileReport:
    """Integrity information for one materialized bundle file."""

    name: str
    path: Path
    size_bytes: int
    sha256: str


@dataclass(frozen=True)
class BundleSpec:
    """Parsed provider-independent bundle specification."""

    bundle_id: str
    dataset_name: str
    release_pin_required: bool
    metadata_file: str
    signal_file: str
    corpus_file: str | None
    id_column: str
    selector_column: str
    selector_values: tuple[str, ...]
    sample_length_aliases: tuple[str, ...]
    channel_count_aliases: tuple[str, ...]
    required_metadata_columns: tuple[str, ...]
    providers: Mapping[str, Mapping[str, Any]]
    expected_sha256: Mapping[str, str]


@dataclass(frozen=True)
class BundleValidation:
    """Validation summary returned to callers and CLI users."""

    spec: BundleSpec
    directory: Path
    metadata_rows: int
    selected_rows: int
    signal_keys: int
    corpus_present: bool
    files: tuple[BundleFileReport, ...]


@dataclass(frozen=True)
class BundleDownload:
    """Provider provenance plus the validated local materialization."""

    provider: str
    requested_revision: str
    directory: Path
    validation: BundleValidation


def load_bundle_spec(bundle_id: str = "cwru-demo-v1") -> BundleSpec:
    """Load a packaged YAML bundle specification by stable bundle id."""
    path = Path(__file__).with_name("manifests") / f"{bundle_id}.yaml"
    if not path.is_file():
        raise FileNotFoundError(f"Unknown PHMFactory bundle: {bundle_id!r}")
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(payload, Mapping):
        raise TypeError(f"Bundle manifest must be a mapping: {path}")

    files = _mapping(payload, "files")
    metadata = _mapping(payload, "metadata")
    selector = _mapping(metadata, "selector")
    aliases = _mapping(metadata, "column_aliases")

    metadata_file = _file_name(files, "metadata", required=True)
    signal_file = _file_name(files, "signals", required=True)
    corpus_file = _file_name(files, "corpus", required=False)
    providers = _mapping(payload, "providers")
    if not providers:
        raise ValueError(f"Bundle {bundle_id!r} declares no providers")

    return BundleSpec(
        bundle_id=str(payload.get("bundle_id") or bundle_id),
        dataset_name=str(payload.get("dataset_name") or ""),
        release_pin_required=bool(payload.get("release_pin_required", True)),
        metadata_file=metadata_file,
        signal_file=signal_file,
        corpus_file=corpus_file,
        id_column=str(metadata.get("id_column") or "Id"),
        selector_column=str(selector.get("column") or "Name"),
        selector_values=tuple(str(value) for value in selector.get("values") or ()),
        sample_length_aliases=tuple(
            str(value) for value in aliases.get("sample_length") or ("Sample_lenth",)
        ),
        channel_count_aliases=tuple(
            str(value) for value in aliases.get("channel_count") or ("Channel",)
        ),
        required_metadata_columns=tuple(
            str(value) for value in metadata.get("required_columns") or ()
        ),
        providers=providers,
        expected_sha256={
            str(name): str(value)
            for name, value in _mapping(payload, "expected_sha256").items()
            if value
        },
    )


def _expected_sha256_for_file(spec: BundleSpec, filename: str) -> str:
    """Resolve one hash pin using provider-neutral logical manifest keys."""
    logical_by_filename = {
        spec.metadata_file: "metadata",
        spec.signal_file: "signals",
    }
    if spec.corpus_file:
        logical_by_filename[spec.corpus_file] = "corpus"

    logical_key = logical_by_filename.get(filename)
    logical_value = str(spec.expected_sha256.get(logical_key, "")) if logical_key else ""
    filename_value = str(spec.expected_sha256.get(filename, ""))
    if (
        logical_value
        and filename_value
        and logical_value.casefold() != filename_value.casefold()
    ):
        raise BundleValidationError(
  f"Conflicting SHA-256 pins for {filename}: "
  f"{logical_key}={logical_value}, {filename}={filename_value}"
        )
    return logical_value or filename_value


def validate_bundle(
    directory: str | Path,
    *,
    spec: BundleSpec | None = None,
) -> BundleValidation:
    """Validate metadata, HDF5 signal keys/shapes, optional corpus, and hashes."""
    bundle_spec = spec or load_bundle_spec()
    root = Path(directory).expanduser().resolve()
    if not root.is_dir():
        raise BundleValidationError(f"Bundle directory does not exist: {root}")

    metadata_path = root / bundle_spec.metadata_file
    signal_path = root / bundle_spec.signal_file
    for required_path in (metadata_path, signal_path):
        if not required_path.is_file():
            raise BundleValidationError(f"Required bundle file is missing: {required_path}")

    headers, metadata_rows = _read_excel_rows(metadata_path)
    _require_column(headers, bundle_spec.id_column, metadata_path)
    _require_column(headers, bundle_spec.selector_column, metadata_path)
    for column in bundle_spec.required_metadata_columns:
        _require_column(headers, column, metadata_path)
    sample_length_column = _find_alias(headers, bundle_spec.sample_length_aliases)
    channel_count_column = _find_alias(headers, bundle_spec.channel_count_aliases)
    if sample_length_column is None:
        raise BundleValidationError(
            f"Metadata is missing a sample-length column; accepted aliases: "
            f"{bundle_spec.sample_length_aliases}"
        )
    if channel_count_column is None:
        raise BundleValidationError(
            f"Metadata is missing a channel-count column; accepted aliases: "
            f"{bundle_spec.channel_count_aliases}"
        )

    selected = [
        row
        for row in metadata_rows
        if not bundle_spec.selector_values
        or str(row.get(bundle_spec.selector_column, "")).strip()
        in bundle_spec.selector_values
    ]
    if not selected:
        raise BundleValidationError(
            f"No rows matched {bundle_spec.selector_column} in "
            f"{bundle_spec.selector_values}"
        )

    selected_id_values = [
        _normalise_id(row[bundle_spec.id_column]) for row in selected
    ]
    duplicate_ids = sorted(
        sample_id
        for sample_id, count in Counter(selected_id_values).items()
        if count > 1
    )
    if duplicate_ids:
        preview = ", ".join(duplicate_ids[:10])
        raise BundleValidationError(
            f"Metadata contains {len(duplicate_ids)} duplicate selected Id value(s): "
            f"{preview}"
        )

    h5py = _import_h5py()
    selected_ids = set(selected_id_values)
    with h5py.File(signal_path, "r") as handle:
        signal_keys = {str(key) for key in handle.keys()}
        missing_ids = sorted(selected_ids - signal_keys)
        if missing_ids:
            preview = ", ".join(missing_ids[:10])
            raise BundleValidationError(
                f"{len(missing_ids)} metadata Id values are absent from "
                f"{signal_path.name}: {preview}"
            )
        for row in selected:
            sample_id = _normalise_id(row[bundle_spec.id_column])
            dataset = handle[sample_id]
            if dataset.ndim != 2:
                raise BundleValidationError(
                    f"HDF5 dataset {sample_id!r} must have shape (L, C); "
                    f"found {dataset.shape}"
                )
            expected_length = _positive_int(row.get(sample_length_column))
            expected_channels = _positive_int(row.get(channel_count_column))
            if expected_length is not None and dataset.shape[0] != expected_length:
                raise BundleValidationError(
                    f"HDF5 dataset {sample_id!r} length {dataset.shape[0]} does not "
                    f"match metadata {expected_length}"
                )
            if expected_channels is not None and dataset.shape[1] != expected_channels:
                raise BundleValidationError(
                    f"HDF5 dataset {sample_id!r} channel count {dataset.shape[1]} "
                    f"does not match metadata {expected_channels}"
                )

    corpus_present = False
    if bundle_spec.corpus_file:
        corpus_path = root / bundle_spec.corpus_file
        if corpus_path.is_file():
            corpus_present = True
            corpus_headers, corpus_rows = _read_excel_rows(corpus_path)
            _require_column(corpus_headers, bundle_spec.id_column, corpus_path)
            corpus_ids = {
                _normalise_id(row[bundle_spec.id_column])
                for row in corpus_rows
                if row.get(bundle_spec.id_column) is not None
            }
            unknown_ids = sorted(corpus_ids - {
                _normalise_id(row[bundle_spec.id_column]) for row in metadata_rows
            })
            if unknown_ids:
                preview = ", ".join(unknown_ids[:10])
                raise BundleValidationError(
                    f"corpus.xlsx references {len(unknown_ids)} unknown Id values: "
                    f"{preview}"
                )

    reports = tuple(
        _file_report(path)
        for path in (
            metadata_path,
            signal_path,
            *((root / bundle_spec.corpus_file,) if corpus_present else ()),
        )
    )
    for report in reports:
        expected = _expected_sha256_for_file(bundle_spec, report.name)
        if expected and report.sha256.lower() != expected.lower():
            raise BundleValidationError(
                f"SHA-256 mismatch for {report.name}: expected {expected}, "
                f"found {report.sha256}"
            )

    return BundleValidation(
        spec=bundle_spec,
        directory=root,
        metadata_rows=len(metadata_rows),
        selected_rows=len(selected),
        signal_keys=len(signal_keys),
        corpus_present=corpus_present,
        files=reports,
    )



def _read_provenance(root: Path) -> Mapping[str, Any]:
    path = root / ".phmfactory-bundle.yaml"
    if not path.is_file():
        return {}
    try:
        payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except (OSError, yaml.YAMLError):
        return {}
    return payload if isinstance(payload, Mapping) else {}


def _provenance_matches(
    root: Path,
    source: str,
    revision: str,
    validation: BundleValidation,
) -> bool:
    payload = _read_provenance(root)
    if payload.get("bundle_id") != validation.spec.bundle_id:
        return False
    if payload.get("provider") != source:
        return False
    if str(payload.get("requested_revision") or "") != revision:
        return False
    if bool(payload.get("corpus_present")) != validation.corpus_present:
        return False
    recorded = payload.get("files")
    if not isinstance(recorded, Mapping):
        return False
    if {str(name) for name in recorded} != {item.name for item in validation.files}:
        return False
    for item in validation.files:
        entry = recorded.get(item.name)
        if not isinstance(entry, Mapping):
            return False
        if str(entry.get("sha256") or "").casefold() != item.sha256.casefold():
            return False
        try:
            if int(entry.get("size_bytes")) != item.size_bytes:
                return False
        except (TypeError, ValueError):
            return False
    return True


def _replace_bundle_files(
    root: Path,
    staging: Path,
    validation: BundleValidation,
) -> None:
    spec = validation.spec
    root.mkdir(parents=True, exist_ok=True)
    staged_names = {report.name for report in validation.files}
    declared = [spec.metadata_file, spec.signal_file]
    if spec.corpus_file:
        declared.append(spec.corpus_file)

    for name in declared:
        target = root / name
        if name not in staged_names:
            if target.is_dir() and not target.is_symlink():
                shutil.rmtree(target)
            elif target.exists() or target.is_symlink():
                target.unlink()
            continue

        source = staging / name
        temp_target = root / f".{name}.phmfactory-tmp"
        if temp_target.exists() or temp_target.is_symlink():
            if temp_target.is_dir() and not temp_target.is_symlink():
                shutil.rmtree(temp_target)
            else:
                temp_target.unlink()
        _materialise(source, temp_target)
        temp_target.replace(target)

    provenance = root / ".phmfactory-bundle.yaml"
    if provenance.exists() or provenance.is_symlink():
        provenance.unlink()

def download_bundle(
    bundle_id: str = "cwru-demo-v1",
    *,
    source: str = "huggingface",
    destination: str | Path | None = None,
    revision: str | None = None,
    force: bool = False,
) -> BundleDownload:
    """Download required bundle files from one explicit public provider."""
    spec = load_bundle_spec(bundle_id)
    if source not in spec.providers:
        available = ", ".join(sorted(spec.providers))
        raise ValueError(f"Unknown data source {source!r}; choose one of: {available}")
    provider = spec.providers[source]
    requested_revision = str(revision or provider.get("revision") or "")
    if not requested_revision:
        raise ValueError(f"Provider {source!r} does not declare a revision")

    root = Path(destination or Path.home() / ".cache" / "phmfactory" / bundle_id)
    root = root.expanduser().resolve()
    root.parent.mkdir(parents=True, exist_ok=True)
    root.mkdir(parents=True, exist_ok=True)

    required_names = (spec.metadata_file, spec.signal_file)
    if not force and all((root / name).is_file() for name in required_names):
        try:
            validation = validate_bundle(root, spec=spec)
        except BundleValidationError:
            validation = None
        if validation is not None and _provenance_matches(
            root,
            source,
            requested_revision,
            validation,
        ):
            return BundleDownload(source, requested_revision, root, validation)

    with tempfile.TemporaryDirectory(
        prefix=f".{bundle_id}-",
        dir=root.parent,
    ) as temp_dir:
        staging = Path(temp_dir)
        if source == "huggingface":
            _download_huggingface(spec, provider, staging, requested_revision)
        elif source == "modelscope":
            _download_modelscope(spec, provider, staging, requested_revision)
        else:  # guarded above; retained for explicit exhaustiveness
            raise ValueError(f"Unsupported provider implementation: {source}")
        staged_validation = validate_bundle(staging, spec=spec)
        _replace_bundle_files(root, staging, staged_validation)

    validation = validate_bundle(root, spec=spec)
    _write_provenance(root, source, requested_revision, validation)
    return BundleDownload(source, requested_revision, root, validation)


def compare_bundle_hashes(
    left: str | Path,
    right: str | Path,
    *,
    spec: BundleSpec | None = None,
) -> Mapping[str, str]:
    """Require identical core-file hashes across two provider materializations."""
    bundle_spec = spec or load_bundle_spec()
    left_report = validate_bundle(left, spec=bundle_spec)
    right_report = validate_bundle(right, spec=bundle_spec)
    left_hashes = {item.name: item.sha256 for item in left_report.files}
    right_hashes = {item.name: item.sha256 for item in right_report.files}

    compared_names = {bundle_spec.metadata_file, bundle_spec.signal_file}
    if left_report.corpus_present or right_report.corpus_present:
        if left_report.corpus_present != right_report.corpus_present:
            raise BundleValidationError(
                "Provider parity failed: corpus.xlsx exists in only one bundle"
            )
        assert bundle_spec.corpus_file is not None
        compared_names.add(bundle_spec.corpus_file)

    mismatches = {
        name: (left_hashes.get(name), right_hashes.get(name))
        for name in sorted(compared_names)
        if left_hashes.get(name) != right_hashes.get(name)
    }
    if mismatches:
        raise BundleValidationError(f"Provider bundle hashes differ: {mismatches}")
    return {name: left_hashes[name] for name in sorted(compared_names)}


def _download_huggingface(
    spec: BundleSpec,
    provider: Mapping[str, Any],
    root: Path,
    revision: str,
) -> None:
    try:
        from huggingface_hub import hf_hub_download
    except ImportError as exc:
        raise ModuleNotFoundError(
            "The Hugging Face provider requires huggingface_hub"
        ) from exc

    repo_id = str(provider.get("repo_id") or "")
    repo_type = str(provider.get("repo_type") or "dataset")
    file_map = _mapping(provider, "files")
    for logical_name, local_name, required in _bundle_files(spec):
        remote_name = str(file_map.get(logical_name) or local_name)
        try:
            downloaded = hf_hub_download(
                repo_id=repo_id,
                filename=remote_name,
                repo_type=repo_type,
                revision=revision,
                local_dir=root,
            )
        except Exception:
            if not required:
                continue
            raise
        _materialise(Path(downloaded), root / local_name)


def _download_modelscope(
    spec: BundleSpec,
    provider: Mapping[str, Any],
    root: Path,
    revision: str,
    *,
    runner: Callable[..., subprocess.CompletedProcess[str]] = subprocess.run,
) -> None:
    executable = shutil.which("modelscope")
    if executable is None:
        raise ModuleNotFoundError(
            "The ModelScope provider requires the modelscope CLI"
        )
    repo_id = str(provider.get("repo_id") or "")
    file_map = _mapping(provider, "files")
    for logical_name, local_name, required in _bundle_files(spec):
        remote_name = str(file_map.get(logical_name) or local_name)
        command = [
            executable,
            "download",
            "--dataset",
            repo_id,
            remote_name,
            "--revision",
            revision,
            "--local_dir",
            str(root),
        ]
        completed = runner(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=900,
        )
        if completed.returncode != 0:
            if not required:
                continue
            raise RuntimeError(
                f"ModelScope failed to download {remote_name!r}: "
                f"{completed.stderr.strip()}"
            )
        candidate = root / remote_name
        if not candidate.is_file():
            matches = list(root.rglob(Path(remote_name).name))
            if len(matches) != 1:
                raise FileNotFoundError(
                    f"ModelScope reported success but {remote_name!r} was not "
                    f"materialized uniquely under {root}"
                )
            candidate = matches[0]
        _materialise(candidate, root / local_name)


def _bundle_files(spec: BundleSpec) -> Sequence[tuple[str, str, bool]]:
    files: list[tuple[str, str, bool]] = [
        ("metadata", spec.metadata_file, True),
        ("signals", spec.signal_file, True),
    ]
    if spec.corpus_file:
        files.append(("corpus", spec.corpus_file, False))
    return tuple(files)


def _write_provenance(
    root: Path,
    source: str,
    revision: str,
    validation: BundleValidation,
) -> None:
    payload = {
        "schema_version": 1,
        "bundle_id": validation.spec.bundle_id,
        "provider": source,
        "requested_revision": revision,
        "corpus_present": validation.corpus_present,
        "files": {
            item.name: {
                "size_bytes": item.size_bytes,
                "sha256": item.sha256,
            }
            for item in validation.files
        },
    }
    (root / ".phmfactory-bundle.yaml").write_text(
        yaml.safe_dump(payload, sort_keys=True),
        encoding="utf-8",
    )


def _read_excel_rows(path: Path) -> tuple[tuple[str, ...], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ModuleNotFoundError(
            "Bundle validation requires openpyxl"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration as exc:
            raise BundleValidationError(f"Excel file is empty: {path}") from exc
        headers = tuple("" if value is None else str(value).strip() for value in raw_headers)
        if len(set(headers)) != len(headers):
            raise BundleValidationError(f"Excel headers are not unique: {path}")
        rows: list[dict[str, Any]] = []
        for values in iterator:
            row = {header: value for header, value in zip(headers, values) if header}
            if any(value is not None for value in row.values()):
                rows.append(row)
        return headers, rows
    finally:
        workbook.close()


def _import_h5py() -> Any:
    try:
        import h5py
    except ImportError as exc:
        raise ModuleNotFoundError("Bundle validation requires h5py") from exc
    return h5py


def _mapping(payload: Mapping[str, Any], key: str) -> Mapping[str, Any]:
    value = payload.get(key) or {}
    if not isinstance(value, Mapping):
        raise TypeError(f"{key} must be a mapping")
    return value


def _file_name(files: Mapping[str, Any], key: str, *, required: bool) -> str | None:
    entry = files.get(key)
    if entry is None:
        if required:
            raise ValueError(f"Bundle manifest is missing files.{key}")
        return None
    if not isinstance(entry, Mapping):
        raise TypeError(f"files.{key} must be a mapping")
    filename = entry.get("filename")
    if not filename:
        if required:
            raise ValueError(f"Bundle manifest is missing files.{key}.filename")
        return None
    return str(filename)


def _require_column(headers: Sequence[str], name: str, path: Path) -> None:
    if name not in headers:
        raise BundleValidationError(f"Required column {name!r} is absent from {path}")


def _find_alias(headers: Sequence[str], aliases: Sequence[str]) -> str | None:
    return next((alias for alias in aliases if alias in headers), None)


def _normalise_id(value: Any) -> str:
    if value is None:
        raise BundleValidationError("Metadata Id cannot be empty")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if not text:
        raise BundleValidationError("Metadata Id cannot be empty")
    return text


def _positive_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        parsed = int(float(value))
    except (TypeError, ValueError) as exc:
        raise BundleValidationError(f"Expected a positive integer, found {value!r}") from exc
    if parsed <= 0:
        raise BundleValidationError(f"Expected a positive integer, found {value!r}")
    return parsed


def _file_report(path: Path) -> BundleFileReport:
    digest = sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return BundleFileReport(
        name=path.name,
        path=path,
        size_bytes=path.stat().st_size,
        sha256=digest.hexdigest(),
    )


def _materialise(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() == target.resolve():
        return
    shutil.copy2(source, target)
