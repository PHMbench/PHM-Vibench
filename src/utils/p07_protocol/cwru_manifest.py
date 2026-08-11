"""Read-only canonical manifest builder for the P07 CWRU protocol.

The protocol intentionally uses only the official 12 kHz drive-end subset for
three fault locations (IR, B, and OR@6), three fault diameters, and four motor
loads.  Building a manifest reads metadata and raw bytes but never writes to
the dataset tree.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Sequence, Tuple


OFFICIAL_SOURCE_URL = (
    "https://engineering.case.edu/bearingdatacenter/"
    "12k-drive-end-bearing-fault-data"
)
DATASET_ID = 1
DATASET_NAME = "RM_001_CWRU"
SUBSET_ID = "cwru-official-12k-de-ir-b-or6-d007-d014-d021-load0-3-v1"
WINDOW_ALGORITHM_ID = "p07-evenly-distributed-nonoverlap-v1"
WINDOW_SIZE = 4096
WINDOW_COUNT = 24


class CWRUManifestError(ValueError):
    """Raised when the fixed CWRU protocol cannot be represented faithfully."""


@dataclass(frozen=True, slots=True)
class OfficialSpecimen:
    """One entry in the fixed official filename/condition registry."""

    file_name: str
    fault_type: str
    label: int
    diameter_code: str
    diameter_mils: int
    fault_level: int
    load_hp: int
    domain_id: int

    @property
    def specimen_key(self) -> str:
        fault_slug = {"IR": "ir", "B": "b", "OR@6": "or6"}[self.fault_type]
        return (
            f"cwru12k-de-{fault_slug}-d{self.diameter_code}-"
            f"load{self.load_hp}"
        )


def _official_specimens() -> Tuple[OfficialSpecimen, ...]:
    blocks = (
        ("007", 7, 1, (("IR", 1, 105), ("B", 2, 118), ("OR@6", 3, 130))),
        ("014", 14, 2, (("IR", 1, 169), ("B", 2, 185), ("OR@6", 3, 197))),
        ("021", 21, 3, (("IR", 1, 209), ("B", 2, 222), ("OR@6", 3, 234))),
    )
    specimens = []
    for diameter_code, diameter_mils, fault_level, faults in blocks:
        for fault_type, label, first_file_number in faults:
            for load_hp in range(4):
                specimens.append(
                    OfficialSpecimen(
                        file_name=f"{first_file_number + load_hp}.mat",
                        fault_type=fault_type,
                        label=label,
                        diameter_code=diameter_code,
                        diameter_mils=diameter_mils,
                        fault_level=fault_level,
                        load_hp=load_hp,
                        domain_id=load_hp,
                    )
                )
    return tuple(sorted(specimens, key=lambda item: item.specimen_key))


OFFICIAL_12K_DRIVE_END_SPECIMENS = _official_specimens()


@dataclass(frozen=True, slots=True)
class WindowCoordinate:
    """Zero-based, half-open coordinate for one deterministic input window."""

    index: int
    start: int
    stop: int

    def to_dict(self) -> Dict[str, int]:
        return {"index": self.index, "start": self.start, "stop": self.stop}


@dataclass(frozen=True, slots=True)
class ManifestSpecimen:
    """Validated metadata and content identity for one selected raw file."""

    specimen_key: str
    metadata_id: int
    file_name: str
    raw_size_bytes: int
    raw_sha256: str
    dataset_id: int
    dataset_name: str
    fault_type: str
    label: int
    diameter_code: str
    diameter_mils: int
    fault_level: int
    domain_id: int
    load_hp: int
    sample_rate_hz: int
    channels: int
    sample_length: int
    file_weight: float
    windows: Tuple[WindowCoordinate, ...]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "specimen_key": self.specimen_key,
            "metadata_id": self.metadata_id,
            "file_name": self.file_name,
            "raw_size_bytes": self.raw_size_bytes,
            "raw_sha256": self.raw_sha256,
            "dataset_id": self.dataset_id,
            "dataset_name": self.dataset_name,
            "fault_type": self.fault_type,
            "label": self.label,
            "diameter_code": self.diameter_code,
            "diameter_mils": self.diameter_mils,
            "fault_level": self.fault_level,
            "domain_id": self.domain_id,
            "load_hp": self.load_hp,
            "sample_rate_hz": self.sample_rate_hz,
            "channels": self.channels,
            "sample_length": self.sample_length,
            "file_weight": self.file_weight,
            "windows": [window.to_dict() for window in self.windows],
        }


@dataclass(frozen=True, slots=True)
class CWRUFold:
    """One load-balanced train/validation/test diameter rotation."""

    fold_id: str
    train_diameter_code: str
    validation_diameter_code: str
    test_diameter_code: str
    train_specimen_keys: Tuple[str, ...]
    validation_specimen_keys: Tuple[str, ...]
    test_specimen_keys: Tuple[str, ...]
    excluded_specimen_keys: Tuple[str, ...]
    evaluation_unit: str = "file"
    weighting: str = "equal_file"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "fold_id": self.fold_id,
            "train_diameter_code": self.train_diameter_code,
            "validation_diameter_code": self.validation_diameter_code,
            "test_diameter_code": self.test_diameter_code,
            "train_specimen_keys": list(self.train_specimen_keys),
            "validation_specimen_keys": list(self.validation_specimen_keys),
            "test_specimen_keys": list(self.test_specimen_keys),
            "excluded_specimen_keys": list(self.excluded_specimen_keys),
            "evaluation_unit": self.evaluation_unit,
            "weighting": self.weighting,
        }


@dataclass(frozen=True, slots=True)
class CWRUManifest:
    """Immutable canonical protocol manifest and its content root."""

    schema_version: int
    subset_id: str
    official_source_url: str
    metadata_subset_sha256: str
    reader_source_sha256: str
    preprocessing_source_sha256: str
    specimens: Tuple[ManifestSpecimen, ...]
    folds: Tuple[CWRUFold, ...]
    root_sha256: str

    def payload(self) -> Dict[str, Any]:
        """Return the canonical payload whose digest is ``root_sha256``."""

        return {
            "schema_version": self.schema_version,
            "subset_id": self.subset_id,
            "official_source_url": self.official_source_url,
            "metadata_subset_sha256": self.metadata_subset_sha256,
            "reader_source_sha256": self.reader_source_sha256,
            "preprocessing_source_sha256": self.preprocessing_source_sha256,
            "dataset_contract": {
                "dataset_id": DATASET_ID,
                "dataset_name": DATASET_NAME,
                "sample_rate_hz": 12000,
                "channels": 2,
                "fault_types": ["IR", "B", "OR@6"],
                "diameter_codes": ["007", "014", "021"],
                "loads_hp": [0, 1, 2, 3],
                "expected_file_count": 36,
            },
            "window_protocol": {
                "algorithm_id": WINDOW_ALGORITHM_ID,
                "coordinate_system": "zero_based_half_open",
                "window_size": WINDOW_SIZE,
                "window_count_per_file": WINDOW_COUNT,
                "overlap_allowed": False,
            },
            "fold_protocol": {
                "fold_count": 3,
                "rotation_axis": "diameter_code",
                "diameter_count_per_split": 1,
                "files_per_split": 12,
                "loads_hp_per_split": [0, 1, 2, 3],
                "fault_types_per_split": ["IR", "B", "OR@6"],
                "all_files_used_per_fold": True,
                "each_file_tested_once_across_folds": True,
                "evaluation_unit": "file",
                "weighting": "equal_file",
            },
            "specimens": [specimen.to_dict() for specimen in self.specimens],
            "folds": [fold.to_dict() for fold in self.folds],
        }

    def to_dict(self) -> Dict[str, Any]:
        result = self.payload()
        result["root_sha256"] = self.root_sha256
        return result

    def canonical_json(self) -> str:
        """Return a portable serialization including the declared root."""

        return _canonical_json(self.to_dict())


def deterministic_window_coordinates(
    sample_length: int,
    *,
    window_size: int = WINDOW_SIZE,
    window_count: int = WINDOW_COUNT,
) -> Tuple[WindowCoordinate, ...]:
    """Distribute fixed windows over a recording without overlap.

    The first window starts at zero and the final window ends at
    ``sample_length``.  Any slack beyond ``window_count * window_size`` is
    distributed monotonically between adjacent windows with integer arithmetic.
    """

    for name, value in {
        "sample_length": sample_length,
        "window_size": window_size,
        "window_count": window_count,
    }.items():
        if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
            raise CWRUManifestError(f"{name} must be a positive integer.")
    if window_count < 2:
        raise CWRUManifestError(
            "window_count must be at least 2 to preserve both recording endpoints."
        )
    required = window_size * window_count
    if sample_length < required:
        raise CWRUManifestError(
            f"sample_length={sample_length} cannot provide {window_count} "
            f"non-overlapping windows of size {window_size}."
        )
    slack = sample_length - required
    denominator = max(window_count - 1, 1)
    windows = tuple(
        WindowCoordinate(
            index=index,
            start=index * window_size + (index * slack) // denominator,
            stop=index * window_size + (index * slack) // denominator + window_size,
        )
        for index in range(window_count)
    )
    if windows[0].start != 0 or windows[-1].stop != sample_length:
        raise AssertionError("Window construction did not preserve endpoint coverage.")
    if any(left.stop > right.start for left, right in zip(windows, windows[1:])):
        raise AssertionError("Window construction produced overlapping intervals.")
    return windows


def build_cwru_manifest(
    *,
    metadata_path: Path,
    raw_dir: Path,
    reader_source_path: Path,
    preprocessing_source_path: Path,
) -> CWRUManifest:
    """Build the fixed P07 CWRU manifest without writing any input path."""

    metadata_path = Path(metadata_path)
    raw_dir = Path(raw_dir)
    reader_source_path = Path(reader_source_path)
    preprocessing_source_path = Path(preprocessing_source_path)
    _require_readable_file(metadata_path, "metadata_path")
    if not raw_dir.is_dir():
        raise CWRUManifestError(f"raw_dir is not a directory: {raw_dir}")
    _require_readable_file(reader_source_path, "reader_source_path")
    _require_readable_file(preprocessing_source_path, "preprocessing_source_path")

    rows = _read_metadata(metadata_path)
    by_file: Dict[str, list[Mapping[str, Any]]] = {}
    expected_names = {item.file_name for item in OFFICIAL_12K_DRIVE_END_SPECIMENS}
    for row in rows:
        file_value = row.get("File")
        if not isinstance(file_value, str) or not file_value.strip():
            continue
        file_name = file_value.strip()
        if file_name in expected_names:
            by_file.setdefault(file_name, []).append(row)

    specimens = []
    metadata_ids = set()
    for expected in OFFICIAL_12K_DRIVE_END_SPECIMENS:
        matches = by_file.get(expected.file_name, [])
        if len(matches) != 1:
            raise CWRUManifestError(
                f"Expected exactly one metadata row for {expected.file_name}, "
                f"found {len(matches)}."
            )
        row = matches[0]
        metadata_id = _finite_int(row.get("Id"), "Id", expected.file_name)
        if metadata_id in metadata_ids:
            raise CWRUManifestError(f"Duplicate metadata Id {metadata_id} in selected subset.")
        metadata_ids.add(metadata_id)
        _require_equal_int(row, "Dataset_id", DATASET_ID, expected.file_name)
        if _text(row.get("Name"), "Name", expected.file_name) != DATASET_NAME:
            raise CWRUManifestError(
                f"{expected.file_name}: Name must be {DATASET_NAME!r}."
            )
        _require_equal_int(row, "Domain_id", expected.domain_id, expected.file_name)
        _require_equal_int(row, "Label", expected.label, expected.file_name)
        _require_equal_int(row, "Fault_level", expected.fault_level, expected.file_name)
        _require_equal_int(row, "Sample_rate", 12000, expected.file_name)
        _require_equal_int(row, "Channel", 2, expected.file_name)
        _validate_optional_load(row, expected)
        sample_length = _finite_int(
            row.get("Sample_lenth"), "Sample_lenth", expected.file_name
        )
        windows = deterministic_window_coordinates(sample_length)

        raw_path = raw_dir / expected.file_name
        _require_readable_file(raw_path, f"raw file {expected.file_name}")
        raw_size = raw_path.stat().st_size
        if raw_size <= 0:
            raise CWRUManifestError(f"Raw file is empty: {expected.file_name}")
        specimens.append(
            ManifestSpecimen(
                specimen_key=expected.specimen_key,
                metadata_id=metadata_id,
                file_name=expected.file_name,
                raw_size_bytes=raw_size,
                raw_sha256=_sha256_file(raw_path),
                dataset_id=DATASET_ID,
                dataset_name=DATASET_NAME,
                fault_type=expected.fault_type,
                label=expected.label,
                diameter_code=expected.diameter_code,
                diameter_mils=expected.diameter_mils,
                fault_level=expected.fault_level,
                domain_id=expected.domain_id,
                load_hp=expected.load_hp,
                sample_rate_hz=12000,
                channels=2,
                sample_length=sample_length,
                file_weight=1.0,
                windows=windows,
            )
        )

    ordered_specimens = tuple(sorted(specimens, key=lambda item: item.specimen_key))
    metadata_payload = [
        {
            key: specimen.to_dict()[key]
            for key in (
                "specimen_key",
                "metadata_id",
                "file_name",
                "dataset_id",
                "dataset_name",
                "fault_type",
                "label",
                "diameter_code",
                "diameter_mils",
                "fault_level",
                "domain_id",
                "load_hp",
                "sample_rate_hz",
                "channels",
                "sample_length",
            )
        }
        for specimen in ordered_specimens
    ]
    metadata_subset_sha256 = _sha256_bytes(_canonical_json(metadata_payload).encode("utf-8"))
    folds = _build_folds(ordered_specimens)
    provisional = CWRUManifest(
        schema_version=1,
        subset_id=SUBSET_ID,
        official_source_url=OFFICIAL_SOURCE_URL,
        metadata_subset_sha256=metadata_subset_sha256,
        reader_source_sha256=_sha256_file(reader_source_path),
        preprocessing_source_sha256=_sha256_file(preprocessing_source_path),
        specimens=ordered_specimens,
        folds=folds,
        root_sha256="",
    )
    root_sha256 = _sha256_bytes(_canonical_json(provisional.payload()).encode("utf-8"))
    return CWRUManifest(
        schema_version=provisional.schema_version,
        subset_id=provisional.subset_id,
        official_source_url=provisional.official_source_url,
        metadata_subset_sha256=provisional.metadata_subset_sha256,
        reader_source_sha256=provisional.reader_source_sha256,
        preprocessing_source_sha256=provisional.preprocessing_source_sha256,
        specimens=provisional.specimens,
        folds=provisional.folds,
        root_sha256=root_sha256,
    )


def _build_folds(specimens: Sequence[ManifestSpecimen]) -> Tuple[CWRUFold, ...]:
    all_keys = {item.specimen_key for item in specimens}
    folds = []
    rotations = (
        ("007", "014", "021"),
        ("014", "021", "007"),
        ("021", "007", "014"),
    )
    for train_diameter, validation_diameter, test_diameter in rotations:
        train = tuple(
            sorted(
                item.specimen_key
                for item in specimens
                if item.diameter_code == train_diameter
            )
        )
        validation = tuple(
            sorted(
                item.specimen_key
                for item in specimens
                if item.diameter_code == validation_diameter
            )
        )
        test = tuple(
            sorted(
                item.specimen_key
                for item in specimens
                if item.diameter_code == test_diameter
            )
        )
        used = set(train) | set(validation) | set(test)
        if len(train) != 12 or len(validation) != 12 or len(test) != 12:
            raise AssertionError("Fixed CWRU fold cardinality is inconsistent.")
        if (set(train) & set(validation)) or (set(train) & set(test)) or (
            set(validation) & set(test)
        ):
            raise AssertionError("Fixed CWRU fold contains overlapping specimen keys.")
        if used != all_keys:
            raise AssertionError("Fixed CWRU fold must use all 36 specimen keys.")
        folds.append(
            CWRUFold(
                fold_id=(
                    f"diameter-train-{train_diameter}-"
                    f"validation-{validation_diameter}-test-{test_diameter}"
                ),
                train_diameter_code=train_diameter,
                validation_diameter_code=validation_diameter,
                test_diameter_code=test_diameter,
                train_specimen_keys=train,
                validation_specimen_keys=validation,
                test_specimen_keys=test,
                excluded_specimen_keys=(),
            )
        )
    for role in (
        "train_specimen_keys",
        "validation_specimen_keys",
        "test_specimen_keys",
    ):
        role_keys = tuple(key for fold in folds for key in getattr(fold, role))
        if len(role_keys) != len(all_keys) or set(role_keys) != all_keys:
            raise AssertionError(
                f"CWRU diameter rotation must assign every specimen once to {role}."
            )
    return tuple(folds)


def _read_metadata(path: Path) -> Tuple[Mapping[str, Any], ...]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                raise CWRUManifestError("Metadata CSV has no header.")
            _validate_metadata_columns(reader.fieldnames)
            return tuple(dict(row) for row in reader)
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as error:  # pragma: no cover - runtime dependency guard
            raise CWRUManifestError("openpyxl is required to read XLSX metadata.") from error
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            iterator = worksheet.iter_rows(values_only=True)
            try:
                header = next(iterator)
            except StopIteration as error:
                raise CWRUManifestError("Metadata workbook is empty.") from error
            fieldnames = tuple(str(value) if value is not None else "" for value in header)
            _validate_metadata_columns(fieldnames)
            return tuple(dict(zip(fieldnames, row)) for row in iterator)
        finally:
            workbook.close()
    raise CWRUManifestError(
        f"Unsupported metadata format {suffix!r}; expected CSV, XLSX, or XLSM."
    )


def _validate_metadata_columns(fieldnames: Iterable[str]) -> None:
    required = {
        "Id",
        "Dataset_id",
        "Name",
        "File",
        "Label",
        "Fault_level",
        "Domain_id",
        "Sample_rate",
        "Sample_lenth",
        "Channel",
    }
    missing = sorted(required.difference(fieldnames))
    if missing:
        raise CWRUManifestError(f"Metadata is missing required columns: {missing}.")


def _validate_optional_load(row: Mapping[str, Any], expected: OfficialSpecimen) -> None:
    for field in ("Load", "Load_hp", "Motor_load"):
        if field in row and row[field] not in (None, ""):
            _require_equal_int(row, field, expected.load_hp, expected.file_name)


def _require_equal_int(
    row: Mapping[str, Any], field: str, expected: int, context: str
) -> None:
    observed = _finite_int(row.get(field), field, context)
    if observed != expected:
        raise CWRUManifestError(
            f"{context}: {field} must be {expected}, observed {observed}."
        )


def _finite_int(value: Any, field: str, context: str) -> int:
    if isinstance(value, bool) or value is None:
        raise CWRUManifestError(f"{context}: {field} must be a finite integer.")
    try:
        numeric = float(value)
    except (TypeError, ValueError) as error:
        raise CWRUManifestError(
            f"{context}: {field} must be a finite integer."
        ) from error
    if not math.isfinite(numeric) or not numeric.is_integer():
        raise CWRUManifestError(f"{context}: {field} must be a finite integer.")
    return int(numeric)


def _text(value: Any, field: str, context: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise CWRUManifestError(f"{context}: {field} must be non-empty text.")
    return value.strip()


def _require_readable_file(path: Path, label: str) -> None:
    if not path.is_file():
        raise CWRUManifestError(f"{label} does not exist or is not a file: {path}")


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
        allow_nan=False,
    )


__all__ = [
    "CWRUFold",
    "CWRUManifest",
    "CWRUManifestError",
    "DATASET_ID",
    "DATASET_NAME",
    "ManifestSpecimen",
    "OFFICIAL_12K_DRIVE_END_SPECIMENS",
    "OFFICIAL_SOURCE_URL",
    "OfficialSpecimen",
    "SUBSET_ID",
    "WINDOW_ALGORITHM_ID",
    "WINDOW_COUNT",
    "WINDOW_SIZE",
    "WindowCoordinate",
    "build_cwru_manifest",
    "deterministic_window_coordinates",
]
