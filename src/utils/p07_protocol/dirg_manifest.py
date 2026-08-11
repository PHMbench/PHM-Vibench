"""Read-only canonical manifest for the P07 DIRG industrial validation.

The local metadata ``Name`` value is the dataset selector.  ``Dataset_id`` is
recorded exactly as observed but is neither a selection key nor assumed stable.
The protocol uses only the balanced C1--C6 operating-condition intersection;
windows remain nested within files and ``Domain_id`` is only a condition-pairing
key, never an authenticated physical-bearing identity.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Final, Iterable, Mapping, Sequence

from . import dirg_preprocessing
from .dirg_preprocessing import WindowCoordinate


SCHEMA_VERSION: Final[int] = 1
PROTOCOL_ID: Final[str] = "P07-DIRG-INDUSTRIAL-VALIDATION-MANIFEST-v1"
SUBSET_ID: Final[str] = "dirg-c1-c6-balanced-domain-severity-rotation-v1"
DATASET_NAME: Final[str] = "RM_020_DIRG"

OFFICIAL_RECORD_ID: Final[int] = 3_559_553
OFFICIAL_RECORD_URL: Final[str] = "https://zenodo.org/records/3559553"
OFFICIAL_API_URL: Final[str] = "https://zenodo.org/api/records/3559553"
DATASET_DOI: Final[str] = "10.5281/zenodo.3559553"
RELATED_ARTICLE_DOI: Final[str] = "10.1016/j.ymssp.2018.10.010"
ACCESS_RIGHT: Final[str] = "open"
LICENSE_ID: Final[str] = "cc-by-4.0"
SOURCE_VERIFIED_ON: Final[str] = "2026-08-01"

CONDITION_IDS: Final[tuple[str, ...]] = ("C1", "C2", "C3", "C4", "C5", "C6")
DOMAIN_IDS: Final[tuple[int, ...]] = (1, 2, 3, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17)
SEVERITY_BY_CONDITION: Final[dict[str, int]] = {
    "C1": 3,
    "C2": 2,
    "C3": 1,
    "C4": 3,
    "C5": 2,
    "C6": 1,
}
CLASS_BY_CONDITION: Final[dict[str, tuple[int, str, int]]] = {
    "C1": (0, "inner_ring", 1),
    "C2": (0, "inner_ring", 1),
    "C3": (0, "inner_ring", 1),
    "C4": (1, "roller", 2),
    "C5": (1, "roller", 2),
    "C6": (1, "roller", 2),
}
ROTATIONS: Final[tuple[tuple[int, int, int], ...]] = (
    (1, 2, 3),
    (2, 3, 1),
    (3, 1, 2),
)

EXPECTED_SAMPLE_RATE_HZ: Final[int] = 51_200
EXPECTED_SAMPLE_LENGTH: Final[int] = 512_000
EXPECTED_CHANNELS: Final[int] = 6
EXPECTED_FILE_COUNT: Final[int] = 78
FILES_PER_SPLIT: Final[int] = 26
WINDOWS_PER_SPLIT: Final[int] = FILES_PER_SPLIT * dirg_preprocessing.WINDOW_COUNT
READER_DOCSTRING_CAVEAT: Final[str] = (
    "docstring_misidentifies_dataset_as_RM_017_Ottawa19"
)
EXCLUSION_REASONS: Final[dict[str, str]] = {
    "exclude_c0_normal": (
        "one healthy C0 condition has no severity 1/2/3 identity and would "
        "cross held-severity splits"
    ),
    "exclude_e4a_endurance": (
        "time-dependent endurance sequence at 102400 Hz, one operating "
        "condition, and 819600 samples"
    ),
    "exclude_incomplete_c1_c6_domains": (
        "Domain_id absent from the complete C1-C6 condition intersection"
    ),
}

_CONDITION_RE = re.compile(r"^(C[0-6])A_")
_SHA256_RE = re.compile(r"[0-9a-f]{64}\Z")
_METADATA_COLUMNS = (
    "Id",
    "Dataset_id",
    "Name",
    "TYPE",
    "File",
    "Label",
    "Label_Description",
    "Fault_level",
    "Domain_id",
    "Domain_description",
    "Sample_rate",
    "Sample_lenth",
    "Channel",
)


class DIRGManifestError(ValueError):
    """Raised when the frozen DIRG protocol cannot be represented exactly."""


@dataclass(frozen=True, slots=True)
class DIRGSpecimen:
    """One selected file with metadata, raw-byte, and coordinate bindings."""

    specimen_key: str
    metadata_id: int
    observed_dataset_id: int
    file_name: str
    metadata_type: str
    condition_id: str
    task_label: int
    class_name: str
    observed_label: int
    label_description: str
    severity: int
    domain_id: int
    domain_description: str
    sample_rate_hz: int
    sample_length: int
    channels: int
    raw_size_bytes: int
    raw_sha256: str
    file_weight: float
    windows: tuple[WindowCoordinate, ...]

    def metadata_payload(self) -> dict[str, Any]:
        return {
            "Id": self.metadata_id,
            "Dataset_id": self.observed_dataset_id,
            "Name": DATASET_NAME,
            "TYPE": self.metadata_type,
            "File": self.file_name,
            "Label": self.observed_label,
            "Label_Description": self.label_description,
            "Fault_level": self.severity,
            "Domain_id": self.domain_id,
            "Domain_description": self.domain_description,
            "Sample_rate": self.sample_rate_hz,
            "Sample_lenth": self.sample_length,
            "Channel": self.channels,
        }

    def to_dict(self) -> dict[str, Any]:
        return {
            "specimen_key": self.specimen_key,
            "metadata": self.metadata_payload(),
            "condition_id": self.condition_id,
            "task_label": self.task_label,
            "class_name": self.class_name,
            "severity": self.severity,
            "raw_size_bytes": self.raw_size_bytes,
            "raw_sha256": self.raw_sha256,
            "file_weight": self.file_weight,
            "windows": [window.to_payload() for window in self.windows],
            "coordinate_set_sha256": dirg_preprocessing.coordinate_set_sha256(
                self.windows
            ),
        }


@dataclass(frozen=True, slots=True)
class DIRGFold:
    """One condition-disjoint held-severity train/validation/test rotation."""

    fold_id: str
    train_severity: int
    validation_severity: int
    test_severity: int
    train_condition_ids: tuple[str, ...]
    validation_condition_ids: tuple[str, ...]
    test_condition_ids: tuple[str, ...]
    train_specimen_keys: tuple[str, ...]
    validation_specimen_keys: tuple[str, ...]
    test_specimen_keys: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "fold_id": self.fold_id,
            "train_severity": self.train_severity,
            "validation_severity": self.validation_severity,
            "test_severity": self.test_severity,
            "train_condition_ids": list(self.train_condition_ids),
            "validation_condition_ids": list(self.validation_condition_ids),
            "test_condition_ids": list(self.test_condition_ids),
            "train_specimen_keys": list(self.train_specimen_keys),
            "validation_specimen_keys": list(self.validation_specimen_keys),
            "test_specimen_keys": list(self.test_specimen_keys),
            "files_per_split": FILES_PER_SPLIT,
            "windows_per_split": WINDOWS_PER_SPLIT,
        }


@dataclass(frozen=True, slots=True)
class ExclusionRecord:
    """Observed rows excluded before any model or score is consulted."""

    code: str
    observed_row_count: int
    file_names: tuple[str, ...]
    reason: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "observed_row_count": self.observed_row_count,
            "file_names": list(self.file_names),
            "reason": self.reason,
        }


@dataclass(frozen=True, slots=True)
class DIRGManifest:
    """Immutable content-rooted DIRG protocol manifest."""

    schema_version: int
    protocol_id: str
    subset_id: str
    metadata_file_sha256: str
    metadata_name_subset_sha256: str
    metadata_selected_subset_sha256: str
    raw_inventory_name_size_sha256: str
    raw_inventory_file_count: int
    raw_unmapped_entries: tuple[str, ...]
    reader_source_sha256: str
    preprocessing_source_sha256: str
    reader_source_caveats: tuple[str, ...]
    observed_dataset_ids: tuple[int, ...]
    specimens: tuple[DIRGSpecimen, ...]
    folds: tuple[DIRGFold, ...]
    exclusions: tuple[ExclusionRecord, ...]
    root_sha256: str

    def payload(self) -> dict[str, Any]:
        preprocessing_contract = dirg_preprocessing.preprocessing_contract_payload()
        preprocessing_contract["source_sha256"] = self.preprocessing_source_sha256
        known_limitations = [
            "physical_bearing_identity_unauthenticated",
            "windows_within_file_are_not_independent_replicates",
        ]
        if READER_DOCSTRING_CAVEAT in self.reader_source_caveats:
            known_limitations.append(
                "local_reader_docstring_misnames_Ottawa19_source_hash_bound"
            )
        return {
            "schema_version": self.schema_version,
            "protocol_id": self.protocol_id,
            "subset_id": self.subset_id,
            "official_source": _official_source_payload(),
            "source_bindings": {
                "metadata_file_sha256": self.metadata_file_sha256,
                "metadata_name_subset_sha256": self.metadata_name_subset_sha256,
                "metadata_selected_subset_sha256": (
                    self.metadata_selected_subset_sha256
                ),
                "raw_inventory_name_size_sha256": (
                    self.raw_inventory_name_size_sha256
                ),
                "raw_inventory_file_count": self.raw_inventory_file_count,
                "raw_unmapped_entries": list(self.raw_unmapped_entries),
                "reader_logical_path": (
                    "src/data_factory/reader/RM_020_DIRG.py"
                ),
                "reader_source_sha256": self.reader_source_sha256,
                "preprocessing_logical_path": (
                    "src/utils/p07_protocol/dirg_preprocessing.py"
                ),
                "preprocessing_source_sha256": self.preprocessing_source_sha256,
                "reader_source_caveats": list(self.reader_source_caveats),
            },
            "dataset_contract": {
                "authoritative_local_name": DATASET_NAME,
                "dataset_selection_key": "Name",
                "observed_dataset_ids": list(self.observed_dataset_ids),
                "dataset_id_semantics": (
                    "observed_only_not_selection_key_not_assumed_stable"
                ),
                "sample_rate_hz": EXPECTED_SAMPLE_RATE_HZ,
                "sample_length": EXPECTED_SAMPLE_LENGTH,
                "sensor_layout": "two_triaxial_accelerometers",
                "channels": EXPECTED_CHANNELS,
                "condition_ids": list(CONDITION_IDS),
                "domain_ids": list(DOMAIN_IDS),
                "domain_id_semantics": (
                    "matched_operating_condition_only_not_independent_bearing"
                ),
                "physical_bearing_identity": "unauthenticated",
                "independent_replicate_unit": "unauthenticated",
                "file_observation_independence_claimed": False,
                "cross_split_physical_bearing_disjointness": "unauditable",
                "expected_file_count": EXPECTED_FILE_COUNT,
                "class_mapping": [
                    {
                        "task_label": 0,
                        "class_name": "inner_ring",
                        "condition_ids": ["C1", "C2", "C3"],
                    },
                    {
                        "task_label": 1,
                        "class_name": "roller",
                        "condition_ids": ["C4", "C5", "C6"],
                    },
                ],
                "requested_outer_nomenclature_status": (
                    "rejected_metadata_authenticates_roller"
                ),
                "severity_by_condition": dict(sorted(SEVERITY_BY_CONDITION.items())),
                "experimental_unit": "file_observation",
                "window_unit": "nested_repeated_observation_not_independent_replicate",
                "file_weighting": "equal_file",
            },
            "window_protocol": {
                **preprocessing_contract,
                "coordinates": [
                    item.to_payload()
                    for item in dirg_preprocessing.uniform_window_coordinates()
                ],
                "overlap_allowed": False,
                "full_record_span_bound": True,
            },
            "fold_protocol": {
                "rotation_axis": "fault_severity",
                "rotations_train_validation_test": [list(item) for item in ROTATIONS],
                "condition_ids_strictly_disjoint_within_fold": True,
                "domain_ids_reused_only_for_operating_condition_matching": True,
                "files_per_split": FILES_PER_SPLIT,
                "windows_per_split": WINDOWS_PER_SPLIT,
                "all_78_files_used_once_per_fold": True,
                "each_file_appears_once_per_role_across_rotations": True,
            },
            "specimens": [item.to_dict() for item in self.specimens],
            "folds": [item.to_dict() for item in self.folds],
            "exclusions": [item.to_dict() for item in self.exclusions],
            "known_limitations": known_limitations,
            "p0_blockers": [
                {
                    "code": "physical_bearing_identity_unauthenticated",
                    "blocks": [
                        "independent_bearing_replication_claim",
                        "cross_split_physical_bearing_disjointness_claim",
                        "independent_bearing_generalization_inference",
                    ],
                    "does_not_block": [
                        "read_only_manifest_validation",
                        "file_level_descriptive_benchmark",
                    ],
                }
            ],
            "claim_boundary": {
                "software_protocol_only": True,
                "evidence_eligible": False,
                "causal_claim_eligible": False,
                "physical_bearing_independence_claimed": False,
            },
        }

    def to_dict(self) -> dict[str, Any]:
        return {**self.payload(), "root_sha256": self.root_sha256}

    def canonical_json(self) -> str:
        return _canonical_json(self.to_dict())


def _official_source_payload() -> dict[str, Any]:
    return {
        "record_id": OFFICIAL_RECORD_ID,
        "official_record_url": OFFICIAL_RECORD_URL,
        "official_api_url": OFFICIAL_API_URL,
        "access_right": ACCESS_RIGHT,
        "license_id": LICENSE_ID,
        "dataset_doi": DATASET_DOI,
        "related_article_doi": RELATED_ARTICLE_DOI,
        "sensor_description": "two_triaxial_accelerometers",
        "expected_sensor_channel_count": EXPECTED_CHANNELS,
        "metadata_verification_basis": "official_zenodo_api_record",
        "metadata_verified_on": SOURCE_VERIFIED_ON,
    }


def build_dirg_manifest(
    *,
    metadata_path: Path,
    raw_dir: Path,
    reader_source_path: Path,
    preprocessing_source_path: Path,
) -> DIRGManifest:
    """Build the fixed read-only manifest from local metadata and raw bytes."""

    metadata_path = Path(metadata_path)
    raw_dir = Path(raw_dir)
    reader_source_path = Path(reader_source_path)
    preprocessing_source_path = Path(preprocessing_source_path)
    _require_file(metadata_path, "metadata_path")
    if not raw_dir.is_dir():
        raise DIRGManifestError(f"raw_dir is not a directory: {raw_dir}")
    _require_file(reader_source_path, "reader_source_path")
    _require_file(preprocessing_source_path, "preprocessing_source_path")

    all_rows = _read_metadata(metadata_path)
    dirg_rows = tuple(
        _normalize_metadata_row(row)
        for row in all_rows
        if row.get("Name") == DATASET_NAME
    )
    if len(dirg_rows) != 180:
        raise DIRGManifestError(
            f"Expected 180 metadata rows selected by Name={DATASET_NAME!r}, "
            f"observed {len(dirg_rows)}."
        )
    _reject_duplicate_rows(dirg_rows)

    c0_rows, standard_rows, endurance_rows = _partition_rows(dirg_rows)
    _validate_excluded_sources(c0_rows, endurance_rows)
    standard_by_condition = _validate_standard_rows(standard_rows)
    common_domains = set.intersection(
        *(set(rows_by_domain) for rows_by_domain in standard_by_condition.values())
    )
    if tuple(sorted(common_domains)) != DOMAIN_IDS:
        raise DIRGManifestError(
            "C1-C6 complete Domain_id intersection drifted: "
            f"{sorted(common_domains)}."
        )
    selected_rows = tuple(
        sorted(
            (
                row
                for row in standard_rows
                if row["Domain_id"] in common_domains
            ),
            key=lambda row: (str(row["File"])[:2], int(row["Domain_id"])),
        )
    )
    incomplete_rows = tuple(
        sorted(
            (row for row in standard_rows if row["Domain_id"] not in common_domains),
            key=lambda row: str(row["File"]),
        )
    )
    if len(selected_rows) != EXPECTED_FILE_COUNT or len(incomplete_rows) != 20:
        raise DIRGManifestError(
            "Balanced DIRG selection must contain 78 selected and 20 "
            "incomplete-domain rows."
        )

    raw_inventory = _raw_inventory(raw_dir)
    metadata_file_names = {str(row["File"]) for row in dirg_rows}
    raw_names = {name for name, _ in raw_inventory}
    missing_raw = sorted(metadata_file_names.difference(raw_names))
    if missing_raw:
        raise DIRGManifestError(f"Metadata files are missing from raw_dir: {missing_raw}.")
    raw_unmapped = tuple(sorted(raw_names.difference(metadata_file_names)))

    coordinates = dirg_preprocessing.uniform_window_coordinates()
    specimens = []
    raw_hashes = set()
    for row in selected_rows:
        file_name = str(row["File"])
        raw_path = raw_dir / file_name
        _require_file(raw_path, f"selected raw file {file_name}")
        if raw_path.is_symlink():
            raise DIRGManifestError(f"Selected raw file cannot be a symlink: {file_name}.")
        raw_size = raw_path.stat().st_size
        if raw_size <= 0:
            raise DIRGManifestError(f"Selected raw file is empty: {file_name}.")
        raw_sha256 = _sha256_file(raw_path)
        if raw_sha256 in raw_hashes:
            raise DIRGManifestError("Selected DIRG files contain duplicate raw bytes.")
        raw_hashes.add(raw_sha256)
        condition_id = _condition_id(file_name)
        task_label, class_name, observed_label = CLASS_BY_CONDITION[condition_id]
        specimen_key = (
            f"dirg-{condition_id.lower()}-domain-{int(row['Domain_id']):02d}"
        )
        specimens.append(
            DIRGSpecimen(
                specimen_key=specimen_key,
                metadata_id=int(row["Id"]),
                observed_dataset_id=int(row["Dataset_id"]),
                file_name=file_name,
                metadata_type=str(row["TYPE"]),
                condition_id=condition_id,
                task_label=task_label,
                class_name=class_name,
                observed_label=observed_label,
                label_description=str(row["Label_Description"]),
                severity=SEVERITY_BY_CONDITION[condition_id],
                domain_id=int(row["Domain_id"]),
                domain_description=str(row["Domain_description"]),
                sample_rate_hz=int(row["Sample_rate"]),
                sample_length=int(row["Sample_lenth"]),
                channels=int(row["Channel"]),
                raw_size_bytes=raw_size,
                raw_sha256=raw_sha256,
                file_weight=1.0,
                windows=coordinates,
            )
        )
    ordered_specimens = tuple(sorted(specimens, key=lambda item: item.specimen_key))
    selected_metadata_payload = [
        item.metadata_payload()
        for item in sorted(ordered_specimens, key=lambda item: item.metadata_id)
    ]
    exclusions = (
        ExclusionRecord(
            code="exclude_c0_normal",
            observed_row_count=len(c0_rows),
            file_names=tuple(sorted(str(row["File"]) for row in c0_rows)),
            reason=EXCLUSION_REASONS["exclude_c0_normal"],
        ),
        ExclusionRecord(
            code="exclude_e4a_endurance",
            observed_row_count=len(endurance_rows),
            file_names=tuple(sorted(str(row["File"]) for row in endurance_rows)),
            reason=EXCLUSION_REASONS["exclude_e4a_endurance"],
        ),
        ExclusionRecord(
            code="exclude_incomplete_c1_c6_domains",
            observed_row_count=len(incomplete_rows),
            file_names=tuple(sorted(str(row["File"]) for row in incomplete_rows)),
            reason=EXCLUSION_REASONS["exclude_incomplete_c1_c6_domains"],
        ),
    )
    reader_bytes = reader_source_path.read_bytes()
    reader_caveats = (
        (READER_DOCSTRING_CAVEAT,)
        if b"RM_017_Ottawa19" in reader_bytes
        else ()
    )
    provisional = DIRGManifest(
        schema_version=SCHEMA_VERSION,
        protocol_id=PROTOCOL_ID,
        subset_id=SUBSET_ID,
        metadata_file_sha256=_sha256_file(metadata_path),
        metadata_name_subset_sha256=_sha256_json(
            sorted(dirg_rows, key=lambda row: int(row["Id"]))
        ),
        metadata_selected_subset_sha256=_sha256_json(selected_metadata_payload),
        raw_inventory_name_size_sha256=_sha256_json(
            [{"name": name, "size_bytes": size} for name, size in raw_inventory]
        ),
        raw_inventory_file_count=len(raw_inventory),
        raw_unmapped_entries=raw_unmapped,
        reader_source_sha256=hashlib.sha256(reader_bytes).hexdigest(),
        preprocessing_source_sha256=_sha256_file(preprocessing_source_path),
        reader_source_caveats=reader_caveats,
        observed_dataset_ids=tuple(
            sorted({item.observed_dataset_id for item in ordered_specimens})
        ),
        specimens=ordered_specimens,
        folds=_build_folds(ordered_specimens),
        exclusions=exclusions,
        root_sha256="",
    )
    rooted = replace(provisional, root_sha256=_sha256_json(provisional.payload()))
    return validate_dirg_manifest(rooted)


def validate_dirg_manifest(manifest: DIRGManifest) -> DIRGManifest:
    """Fail closed on content root, selection, coordinates, or fold leakage."""

    if not isinstance(manifest, DIRGManifest):
        raise TypeError("manifest must be a DIRGManifest.")
    if (
        manifest.schema_version != SCHEMA_VERSION
        or manifest.protocol_id != PROTOCOL_ID
        or manifest.subset_id != SUBSET_ID
    ):
        raise DIRGManifestError("DIRG manifest schema/protocol identity drifted.")
    for label, value in (
        ("metadata_file_sha256", manifest.metadata_file_sha256),
        ("metadata_name_subset_sha256", manifest.metadata_name_subset_sha256),
        ("metadata_selected_subset_sha256", manifest.metadata_selected_subset_sha256),
        ("raw_inventory_name_size_sha256", manifest.raw_inventory_name_size_sha256),
        ("reader_source_sha256", manifest.reader_source_sha256),
        ("preprocessing_source_sha256", manifest.preprocessing_source_sha256),
        ("root_sha256", manifest.root_sha256),
    ):
        _require_sha256(value, label)
    if len(manifest.specimens) != EXPECTED_FILE_COUNT:
        raise DIRGManifestError("DIRG manifest must contain exactly 78 files.")
    keys = [item.specimen_key for item in manifest.specimens]
    names = [item.file_name for item in manifest.specimens]
    metadata_ids = [item.metadata_id for item in manifest.specimens]
    raw_hashes = [item.raw_sha256 for item in manifest.specimens]
    if any(
        len(values) != len(set(values))
        for values in (keys, names, metadata_ids, raw_hashes)
    ):
        raise DIRGManifestError("DIRG manifest contains duplicate file identities.")
    expected_cells = {
        (condition_id, domain_id)
        for condition_id in CONDITION_IDS
        for domain_id in DOMAIN_IDS
    }
    observed_cells = set()
    expected_coordinates = dirg_preprocessing.uniform_window_coordinates()
    for item in manifest.specimens:
        if not isinstance(item, DIRGSpecimen):
            raise TypeError("manifest specimens must be DIRGSpecimen objects.")
        if item.condition_id not in CONDITION_IDS:
            raise DIRGManifestError("DIRG specimen condition is not C1-C6.")
        task_label, class_name, observed_label = CLASS_BY_CONDITION[item.condition_id]
        if (
            item.task_label != task_label
            or item.class_name != class_name
            or item.observed_label != observed_label
            or item.severity != SEVERITY_BY_CONDITION[item.condition_id]
        ):
            raise DIRGManifestError("DIRG class/severity mapping drifted.")
        if (
            _condition_id(item.file_name) != item.condition_id
            or item.specimen_key
            != f"dirg-{item.condition_id.lower()}-domain-{item.domain_id:02d}"
        ):
            raise DIRGManifestError("DIRG filename/specimen identity binding drifted.")
        required_term = (
            "inner ring defect" if item.class_name == "inner_ring" else "roller defect"
        )
        if required_term not in item.label_description.lower():
            raise DIRGManifestError("DIRG metadata class description binding drifted.")
        if (
            item.sample_rate_hz != EXPECTED_SAMPLE_RATE_HZ
            or item.sample_length != EXPECTED_SAMPLE_LENGTH
            or item.channels != EXPECTED_CHANNELS
            or item.metadata_type != "Vibration"
        ):
            raise DIRGManifestError("DIRG selected reader contract drifted.")
        if item.raw_size_bytes <= 0 or item.file_weight != 1.0:
            raise DIRGManifestError("DIRG raw size or equal-file weighting drifted.")
        _require_sha256(item.raw_sha256, f"raw hash for {item.file_name}")
        if item.windows != expected_coordinates:
            raise DIRGManifestError("DIRG specimen coordinates drifted.")
        observed_cells.add((item.condition_id, item.domain_id))
    if observed_cells != expected_cells:
        raise DIRGManifestError("DIRG manifest is not the complete 6x13 condition grid.")
    if tuple(sorted({item.domain_id for item in manifest.specimens})) != DOMAIN_IDS:
        raise DIRGManifestError("DIRG Domain_id intersection drifted.")
    if manifest.observed_dataset_ids != tuple(
        sorted({item.observed_dataset_id for item in manifest.specimens})
    ):
        raise DIRGManifestError("Observed Dataset_id recording drifted.")
    selected_payload = [
        item.metadata_payload()
        for item in sorted(manifest.specimens, key=lambda item: item.metadata_id)
    ]
    if manifest.metadata_selected_subset_sha256 != _sha256_json(selected_payload):
        raise DIRGManifestError("Selected metadata subset hash drifted.")
    for domain_id in DOMAIN_IDS:
        descriptions = {
            item.domain_description
            for item in manifest.specimens
            if item.domain_id == domain_id
        }
        if len(descriptions) != 1:
            raise DIRGManifestError(
                "Domain_id operating-condition descriptions do not match across C1-C6."
            )
    expected_folds = _build_folds(manifest.specimens)
    if manifest.folds != expected_folds:
        raise DIRGManifestError("DIRG fold layout or condition isolation drifted.")
    expected_exclusion_counts = {
        "exclude_c0_normal": 17,
        "exclude_e4a_endurance": 65,
        "exclude_incomplete_c1_c6_domains": 20,
    }
    if len(manifest.exclusions) != len(expected_exclusion_counts) or {
        item.code: item.observed_row_count for item in manifest.exclusions
    } != expected_exclusion_counts:
        raise DIRGManifestError("DIRG exclusion inventory drifted.")
    excluded_names = []
    for item in manifest.exclusions:
        if (
            item.reason != EXCLUSION_REASONS[item.code]
            or item.observed_row_count != len(item.file_names)
            or tuple(sorted(item.file_names)) != item.file_names
        ):
            raise DIRGManifestError("DIRG exclusion reason or file binding drifted.")
        excluded_names.extend(item.file_names)
    if (
        len(excluded_names) != 102
        or len(set(excluded_names)) != len(excluded_names)
        or set(excluded_names).intersection(names)
    ):
        raise DIRGManifestError("DIRG selected and excluded file identities overlap or drifted.")
    if (
        tuple(sorted(manifest.raw_unmapped_entries)) != manifest.raw_unmapped_entries
        or len(manifest.raw_unmapped_entries) != len(set(manifest.raw_unmapped_entries))
        or set(manifest.raw_unmapped_entries).intersection(names)
        or manifest.raw_inventory_file_count
        != 180 + len(manifest.raw_unmapped_entries)
    ):
        raise DIRGManifestError("DIRG raw inventory/unmapped-file binding drifted.")
    if manifest.reader_source_caveats not in {(), (READER_DOCSTRING_CAVEAT,)}:
        raise DIRGManifestError("DIRG reader caveat inventory drifted.")
    expected_root = _sha256_json(manifest.payload())
    if manifest.root_sha256 != expected_root:
        raise DIRGManifestError("DIRG manifest root SHA-256 mismatch.")
    return manifest


def verify_dirg_source_bindings(
    manifest: DIRGManifest,
    *,
    metadata_path: Path,
    raw_dir: Path,
    reader_source_path: Path,
    preprocessing_source_path: Path,
) -> DIRGManifest:
    """Re-read bound bytes and reject any local source or raw-file drift."""

    validate_dirg_manifest(manifest)
    paths = {
        "metadata_file_sha256": Path(metadata_path),
        "reader_source_sha256": Path(reader_source_path),
        "preprocessing_source_sha256": Path(preprocessing_source_path),
    }
    for attribute, path in paths.items():
        _require_file(path, attribute)
        if _sha256_file(path) != getattr(manifest, attribute):
            raise DIRGManifestError(f"DIRG source binding drifted: {attribute}.")
    raw_dir = Path(raw_dir)
    if not raw_dir.is_dir():
        raise DIRGManifestError(f"raw_dir is not a directory: {raw_dir}")
    inventory = _raw_inventory(raw_dir)
    inventory_hash = _sha256_json(
        [{"name": name, "size_bytes": size} for name, size in inventory]
    )
    if (
        len(inventory) != manifest.raw_inventory_file_count
        or inventory_hash != manifest.raw_inventory_name_size_sha256
    ):
        raise DIRGManifestError("DIRG raw directory name/size inventory drifted.")
    for item in manifest.specimens:
        path = raw_dir / item.file_name
        _require_file(path, f"selected raw file {item.file_name}")
        if path.stat().st_size != item.raw_size_bytes or _sha256_file(path) != item.raw_sha256:
            raise DIRGManifestError(f"DIRG raw-byte binding drifted: {item.file_name}.")
    return manifest


def strict_canonical_json_loads(serialized: str | bytes) -> Any:
    """Parse finite duplicate-free JSON and require its exact canonical bytes."""

    if isinstance(serialized, str):
        raw = serialized.encode("utf-8")
    elif isinstance(serialized, bytes):
        raw = serialized
    else:
        raise TypeError("serialized manifest must be str or bytes.")
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as error:
        raise DIRGManifestError("Serialized manifest is not UTF-8.") from error

    def object_pairs(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            if key in result:
                raise DIRGManifestError(f"Duplicate JSON key {key!r}.")
            result[key] = value
        return result

    def invalid_constant(value: str) -> Any:
        raise DIRGManifestError(f"Non-finite JSON constant {value!r} is forbidden.")

    try:
        parsed = json.loads(
            text,
            object_pairs_hook=object_pairs,
            parse_constant=invalid_constant,
        )
    except (json.JSONDecodeError, UnicodeDecodeError) as error:
        raise DIRGManifestError("Serialized manifest is invalid JSON.") from error
    _reject_nonfinite_json(parsed)
    if _canonical_json(parsed).encode("utf-8") != raw:
        raise DIRGManifestError("Serialized manifest is not canonical JSON.")
    return parsed


def validate_serialized_dirg_manifest(
    serialized: str | bytes,
    *,
    expected_manifest: DIRGManifest | None = None,
) -> Mapping[str, Any]:
    """Validate canonical form, declared self-hash, and optional exact manifest."""

    parsed = strict_canonical_json_loads(serialized)
    if not isinstance(parsed, dict):
        raise DIRGManifestError("Serialized DIRG manifest must be a JSON object.")
    expected_keys = {
        "schema_version",
        "protocol_id",
        "subset_id",
        "official_source",
        "source_bindings",
        "dataset_contract",
        "window_protocol",
        "fold_protocol",
        "specimens",
        "folds",
        "exclusions",
        "known_limitations",
        "p0_blockers",
        "claim_boundary",
        "root_sha256",
    }
    if set(parsed) != expected_keys:
        raise DIRGManifestError("Serialized DIRG manifest top-level schema drifted.")
    if parsed.get("protocol_id") != PROTOCOL_ID or parsed.get("subset_id") != SUBSET_ID:
        raise DIRGManifestError("Serialized DIRG protocol identity drifted.")
    declared_root = parsed.get("root_sha256")
    _require_sha256(declared_root, "serialized root_sha256")
    payload = dict(parsed)
    payload.pop("root_sha256")
    if _sha256_json(payload) != declared_root:
        raise DIRGManifestError("Serialized DIRG manifest self-hash mismatch.")
    if expected_manifest is not None and parsed != expected_manifest.to_dict():
        raise DIRGManifestError("Serialized DIRG manifest differs from expected manifest.")
    return parsed


def _build_folds(specimens: Sequence[DIRGSpecimen]) -> tuple[DIRGFold, ...]:
    all_keys = {item.specimen_key for item in specimens}
    folds = []
    for train_severity, validation_severity, test_severity in ROTATIONS:
        conditions_by_role = {
            "train": tuple(
                sorted(
                    condition
                    for condition, severity in SEVERITY_BY_CONDITION.items()
                    if severity == train_severity
                )
            ),
            "validation": tuple(
                sorted(
                    condition
                    for condition, severity in SEVERITY_BY_CONDITION.items()
                    if severity == validation_severity
                )
            ),
            "test": tuple(
                sorted(
                    condition
                    for condition, severity in SEVERITY_BY_CONDITION.items()
                    if severity == test_severity
                )
            ),
        }
        if any(
            set(conditions_by_role[left]).intersection(conditions_by_role[right])
            for left, right in (
                ("train", "validation"),
                ("train", "test"),
                ("validation", "test"),
            )
        ):
            raise DIRGManifestError("DIRG fold condition IDs overlap across splits.")

        def keys_for(role: str) -> tuple[str, ...]:
            return tuple(
                sorted(
                    item.specimen_key
                    for item in specimens
                    if item.condition_id in conditions_by_role[role]
                )
            )

        train = keys_for("train")
        validation = keys_for("validation")
        test = keys_for("test")
        if tuple(map(len, (train, validation, test))) != (
            FILES_PER_SPLIT,
            FILES_PER_SPLIT,
            FILES_PER_SPLIT,
        ):
            raise DIRGManifestError("DIRG fold split does not contain exactly 26 files.")
        if set(train) & set(validation) or set(train) & set(test) or set(validation) & set(test):
            raise DIRGManifestError("DIRG fold file identities overlap across splits.")
        if set(train) | set(validation) | set(test) != all_keys:
            raise DIRGManifestError("DIRG fold does not use all 78 selected files.")
        folds.append(
            DIRGFold(
                fold_id=(
                    f"severity-train-{train_severity}-validation-"
                    f"{validation_severity}-test-{test_severity}"
                ),
                train_severity=train_severity,
                validation_severity=validation_severity,
                test_severity=test_severity,
                train_condition_ids=conditions_by_role["train"],
                validation_condition_ids=conditions_by_role["validation"],
                test_condition_ids=conditions_by_role["test"],
                train_specimen_keys=train,
                validation_specimen_keys=validation,
                test_specimen_keys=test,
            )
        )
    for role in (
        "train_specimen_keys",
        "validation_specimen_keys",
        "test_specimen_keys",
    ):
        assigned = tuple(key for fold in folds for key in getattr(fold, role))
        if len(assigned) != len(all_keys) or set(assigned) != all_keys:
            raise DIRGManifestError(
                f"Each DIRG file must appear exactly once in role {role!r}."
            )
    return tuple(folds)


def _partition_rows(
    rows: Sequence[Mapping[str, Any]],
) -> tuple[
    tuple[Mapping[str, Any], ...],
    tuple[Mapping[str, Any], ...],
    tuple[Mapping[str, Any], ...],
]:
    c0 = []
    standard = []
    endurance = []
    for row in rows:
        file_name = str(row["File"])
        condition_match = _CONDITION_RE.match(file_name)
        if condition_match and condition_match.group(1) == "C0":
            c0.append(row)
        elif condition_match and condition_match.group(1) in CONDITION_IDS:
            standard.append(row)
        elif file_name.startswith("E4A"):
            endurance.append(row)
        else:
            raise DIRGManifestError(f"Unregistered DIRG metadata file family: {file_name}.")
    if tuple(map(len, (c0, standard, endurance))) != (17, 98, 65):
        raise DIRGManifestError("DIRG metadata family counts drifted from 17/98/65.")
    return tuple(c0), tuple(standard), tuple(endurance)


def _validate_excluded_sources(
    c0_rows: Sequence[Mapping[str, Any]],
    endurance_rows: Sequence[Mapping[str, Any]],
) -> None:
    if any(
        row["Fault_level"] != 0
        or row["Sample_rate"] != EXPECTED_SAMPLE_RATE_HZ
        or row["Sample_lenth"] != EXPECTED_SAMPLE_LENGTH
        or row["Channel"] != EXPECTED_CHANNELS
        for row in c0_rows
    ):
        raise DIRGManifestError("C0 exclusion source metadata drifted.")
    if any(
        row["Sample_rate"] != 102_400
        or row["Sample_lenth"] != 819_600
        or row["Channel"] != EXPECTED_CHANNELS
        or row["Domain_id"] != 12
        for row in endurance_rows
    ):
        raise DIRGManifestError("E4A endurance exclusion source metadata drifted.")


def _validate_standard_rows(
    rows: Sequence[Mapping[str, Any]],
) -> dict[str, dict[int, Mapping[str, Any]]]:
    by_condition: dict[str, dict[int, Mapping[str, Any]]] = {
        condition: {} for condition in CONDITION_IDS
    }
    for row in rows:
        file_name = str(row["File"])
        condition = _condition_id(file_name)
        severity = SEVERITY_BY_CONDITION[condition]
        task_label, class_name, observed_label = CLASS_BY_CONDITION[condition]
        del task_label
        if (
            row["Fault_level"] != severity
            or row["Label"] != observed_label
            or row["Sample_rate"] != EXPECTED_SAMPLE_RATE_HZ
            or row["Sample_lenth"] != EXPECTED_SAMPLE_LENGTH
            or row["Channel"] != EXPECTED_CHANNELS
            or row["TYPE"] != "Vibration"
        ):
            raise DIRGManifestError(f"Selected metadata contract drifted: {file_name}.")
        description = str(row["Label_Description"]).lower()
        required_term = "inner ring defect" if class_name == "inner_ring" else "roller defect"
        if required_term not in description:
            raise DIRGManifestError(
                f"{file_name}: metadata does not authenticate class {class_name!r}."
            )
        domain_id = int(row["Domain_id"])
        if domain_id in by_condition[condition]:
            raise DIRGManifestError(
                f"Duplicate condition/domain cell {condition}/{domain_id}."
            )
        by_condition[condition][domain_id] = row
    if set(by_condition) != set(CONDITION_IDS) or any(
        not values for values in by_condition.values()
    ):
        raise DIRGManifestError("DIRG C1-C6 condition registry is incomplete.")
    return by_condition


def _condition_id(file_name: str) -> str:
    match = _CONDITION_RE.match(file_name)
    if match is None or match.group(1) not in CONDITION_IDS:
        raise DIRGManifestError(f"Cannot derive C1-C6 condition from {file_name!r}.")
    return match.group(1)


def _normalize_metadata_row(row: Mapping[str, Any]) -> dict[str, Any]:
    file_name = _text(row.get("File"), "File", DATASET_NAME)
    normalized = {
        "Id": _finite_int(row.get("Id"), "Id", file_name),
        "Dataset_id": _finite_int(row.get("Dataset_id"), "Dataset_id", file_name),
        "Name": _text(row.get("Name"), "Name", file_name),
        "TYPE": _text(row.get("TYPE"), "TYPE", file_name),
        "File": file_name,
        "Label": _finite_int(row.get("Label"), "Label", file_name),
        "Label_Description": _text(
            row.get("Label_Description"), "Label_Description", file_name
        ),
        "Fault_level": _finite_int(row.get("Fault_level"), "Fault_level", file_name),
        "Domain_id": _finite_int(row.get("Domain_id"), "Domain_id", file_name),
        "Domain_description": _text(
            row.get("Domain_description"), "Domain_description", file_name
        ),
        "Sample_rate": _finite_int(row.get("Sample_rate"), "Sample_rate", file_name),
        "Sample_lenth": _finite_int(
            row.get("Sample_lenth"), "Sample_lenth", file_name
        ),
        "Channel": _finite_int(row.get("Channel"), "Channel", file_name),
    }
    if normalized["Name"] != DATASET_NAME:
        raise DIRGManifestError(f"{file_name}: Name must be {DATASET_NAME!r}.")
    return normalized


def _reject_duplicate_rows(rows: Sequence[Mapping[str, Any]]) -> None:
    for field in ("Id", "File"):
        values = [row[field] for row in rows]
        if len(values) != len(set(values)):
            raise DIRGManifestError(f"DIRG metadata contains duplicate {field} values.")


def _read_metadata(path: Path) -> tuple[Mapping[str, Any], ...]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                raise DIRGManifestError("Metadata CSV has no header.")
            _validate_metadata_columns(reader.fieldnames)
            return tuple(dict(row) for row in reader)
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as error:  # pragma: no cover
            raise DIRGManifestError("openpyxl is required for XLSX metadata.") from error
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            iterator = worksheet.iter_rows(values_only=True)
            try:
                header = next(iterator)
            except StopIteration as error:
                raise DIRGManifestError("Metadata workbook is empty.") from error
            fieldnames = tuple(str(value) if value is not None else "" for value in header)
            _validate_metadata_columns(fieldnames)
            return tuple(dict(zip(fieldnames, row)) for row in iterator)
        finally:
            workbook.close()
    raise DIRGManifestError(f"Unsupported metadata format {suffix!r}.")


def _validate_metadata_columns(fieldnames: Iterable[str]) -> None:
    missing = sorted(set(_METADATA_COLUMNS).difference(fieldnames))
    if missing:
        raise DIRGManifestError(f"Metadata is missing required columns: {missing}.")


def _raw_inventory(raw_dir: Path) -> tuple[tuple[str, int], ...]:
    inventory = []
    for path in raw_dir.iterdir():
        if path.is_file():
            inventory.append((path.name, path.stat().st_size))
    return tuple(sorted(inventory))


def _finite_int(value: Any, field: str, context: str) -> int:
    if value is None or isinstance(value, bool):
        raise DIRGManifestError(f"{context}: {field} must be a finite integer.")
    try:
        numeric = float(value)
    except (TypeError, ValueError) as error:
        raise DIRGManifestError(
            f"{context}: {field} must be a finite integer."
        ) from error
    if not math.isfinite(numeric) or not numeric.is_integer():
        raise DIRGManifestError(f"{context}: {field} must be a finite integer.")
    return int(numeric)


def _text(value: Any, field: str, context: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise DIRGManifestError(f"{context}: {field} must be non-empty text.")
    return value.strip()


def _require_file(path: Path, label: str) -> None:
    if not path.is_file():
        raise DIRGManifestError(f"{label} does not exist or is not a file: {path}")


def _require_sha256(value: Any, label: str) -> str:
    if not isinstance(value, str) or _SHA256_RE.fullmatch(value) is None:
        raise DIRGManifestError(f"{label} must be a lowercase SHA-256 hex string.")
    return value


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha256_json(value: Any) -> str:
    return hashlib.sha256(_canonical_json(value).encode("utf-8")).hexdigest()


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
        allow_nan=False,
    )


def _reject_nonfinite_json(value: Any) -> None:
    if isinstance(value, float) and not math.isfinite(value):
        raise DIRGManifestError("Canonical manifest cannot contain non-finite numbers.")
    if isinstance(value, dict):
        for child in value.values():
            _reject_nonfinite_json(child)
    elif isinstance(value, list):
        for child in value:
            _reject_nonfinite_json(child)


__all__ = [
    "ACCESS_RIGHT",
    "CLASS_BY_CONDITION",
    "CONDITION_IDS",
    "DATASET_DOI",
    "DATASET_NAME",
    "DIRGFold",
    "DIRGManifest",
    "DIRGManifestError",
    "DIRGSpecimen",
    "DOMAIN_IDS",
    "EXPECTED_CHANNELS",
    "EXPECTED_FILE_COUNT",
    "EXPECTED_SAMPLE_LENGTH",
    "EXPECTED_SAMPLE_RATE_HZ",
    "ExclusionRecord",
    "FILES_PER_SPLIT",
    "LICENSE_ID",
    "OFFICIAL_API_URL",
    "OFFICIAL_RECORD_ID",
    "OFFICIAL_RECORD_URL",
    "PROTOCOL_ID",
    "RELATED_ARTICLE_DOI",
    "ROTATIONS",
    "SCHEMA_VERSION",
    "SEVERITY_BY_CONDITION",
    "SOURCE_VERIFIED_ON",
    "SUBSET_ID",
    "WINDOWS_PER_SPLIT",
    "build_dirg_manifest",
    "strict_canonical_json_loads",
    "validate_dirg_manifest",
    "validate_serialized_dirg_manifest",
    "verify_dirg_source_bindings",
]
